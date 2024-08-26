import pywinauto
import re
import keyboard
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
import time

# Define the regex pattern for the phone number, including Persian digits
phone_pattern = re.compile(r"(0[0-9]{10}|۰[۰-۹]{10})")

# Excel file name
excel_file = 'phone_numbers.xlsx'

# Check if Excel file exists, if not create one
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.append(['Phone Number', 'Date', 'Time', 'Application/Tab Title'])
    wb.save(excel_file)

# Load the workbook and active sheet
wb = load_workbook(excel_file)
ws = wb.active

# Dictionary to store phone numbers and associated data
phone_data = {}

def normalize_phone_number(phone_number):
    # Map Persian digits to Latin digits
    persian_to_latin = str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789")
    # Normalize by translating Persian digits to Latin digits
    normalized_number = phone_number.translate(persian_to_latin)
    return normalized_number

def save_to_excel():
    for phone_number, entries in phone_data.items():
        # Check if phone number already exists in the Excel sheet
        found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == phone_number:
                # If found, merge the data
                ws[f"B{row[0]}"].value += "\n" + "\n".join(entries[0])
                ws[f"C{row[0]}"].value += "\n" + "\n".join(entries[1])
                ws[f"D{row[0]}"].value += "\n" + "\n".join(entries[2])
                found = True
                break

        if not found:
            # If not found, add a new row
            ws.append([phone_number, "\n".join(entries[0]), "\n".join(entries[1]), "\n".join(entries[2])])

    wb.save(excel_file)

def get_browser_tab_title(window_title):
    # Detect common browser titles and strip unnecessary parts
    if " - Google Chrome" in window_title:
        return window_title.replace(" - Google Chrome", "")
    elif " - Mozilla Firefox" in window_title:
        return window_title.replace(" - Mozilla Firefox", "")
    elif " - Microsoft Edge" in window_title:
        return window_title.replace(" - Microsoft Edge", "")
    elif " - Internet Explorer" in window_title:
        return window_title.replace(" - Internet Explorer", "")
    else:
        return window_title

def monitor_edit_fields():
    app = pywinauto.Application().connect(path="explorer.exe")
    while True:
        try:
            active_window = app.top_window()
            active_window_title = active_window.window_text()

            # Check if the active window is a browser tab
            if any(browser in active_window_title for browser in ["Google Chrome", "Mozilla Firefox", "Microsoft Edge", "Internet Explorer"]):
                app_name = get_browser_tab_title(active_window_title)
            else:
                app_name = active_window_title

            for edit in active_window.descendants(control_type="Edit"):
                text = edit.window_text()
                match = phone_pattern.search(text)
                if match:
                    phone_number = normalize_phone_number(match.group(0))
                    now = datetime.now()
                    date = now.strftime("%Y-%m-%d")
                    time_str = now.strftime("%H:%M:%S")
                    
                    # If the phone number is already in the dictionary, merge the data
                    if phone_number in phone_data:
                        phone_data[phone_number][0].append(date)
                        phone_data[phone_number][1].append(time_str)
                        phone_data[phone_number][2].append(app_name)
                    else:
                        phone_data[phone_number] = [[date], [time_str], [app_name]]
                    
                    print(f"Captured: {phone_number} from {app_name} at {date} {time_str}")

        except Exception as e:
            print(f"Error: {str(e)}")
        time.sleep(1)  # Adjust the sleep interval as needed

def on_finalization_event(e):
    # Triggered when Enter or Tab key is pressed, meaning the input is likely finalized
    monitor_edit_fields()
    save_to_excel()  # Save data whenever a key event finalizes

if __name__ == "__main__":
    # Set up keyboard hooks to detect when Enter or Tab is pressed
    keyboard.on_press_key("enter", on_finalization_event)
    keyboard.on_press_key("tab", on_finalization_event)
    
    # Keep the script running
    keyboard.wait("esc")  # Press 'Esc' to stop the script
